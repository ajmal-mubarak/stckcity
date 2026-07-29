from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def generate_orders_excel(queryset):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Orders"

    headers = [
        "Order Number",
        "Shop",
        "Status",
        "Required Date",
        "Total Amount",
        "Created At",
    ]

    header_fill = PatternFill(
        fill_type="solid",
        start_color="0D6EFD",
    )

    bold = Font(
        bold=True,
        color="FFFFFF",
    )

    for column, header in enumerate(headers, start=1):

        cell = sheet.cell(
            row=1,
            column=column,
            value=header,
        )

        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    row = 2

    for order in queryset:

        sheet.cell(
            row=row,
            column=1,
            value=order.order_number,
        )

        sheet.cell(
            row=row,
            column=2,
            value=order.shop_name,
        )

        sheet.cell(
            row=row,
            column=3,
            value=order.status,
        )

        sheet.cell(
            row=row,
            column=4,
            value=str(order.required_date),
        )

        sheet.cell(
            row=row,
            column=5,
            value=float(order.total_amount),
        )

        sheet.cell(
            row=row,
            column=6,
            value=order.created_at.strftime("%d-%m-%Y %H:%M"),
        )

        row += 1

    return workbook

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def generate_shop_history_excel(shop, orders):

    workbook = Workbook()

    sheet = workbook.active

    sheet.title = "Shop History"

    sheet["A1"] = "SHOP HISTORY REPORT"
    sheet["A1"].font = Font(
        bold=True,
        size=18,
    )

    sheet["A3"] = "Shop Name"
    sheet["B3"] = shop.shop_name

    sheet["A4"] = "Owner"
    sheet["B4"] = shop.owner_name

    sheet["A5"] = "Mobile"
    sheet["B5"] = shop.user.mobile_number

    sheet["A6"] = "Place"
    sheet["B6"] = shop.place

    headers = [
        "Order Number",
        "Date",
        "Status",
        "Required Date",
        "Total",
    ]

    fill = PatternFill(
        fill_type="solid",
        start_color="0D6EFD",
    )

    font = Font(
        bold=True,
        color="FFFFFF",
    )

    row = 8

    for column, header in enumerate(headers, start=1):

        cell = sheet.cell(
            row=row,
            column=column,
            value=header,
        )

        cell.fill = fill
        cell.font = font

    row += 1

    total_purchase = 0

    for order in orders:

        sheet.cell(row=row, column=1).value = order.order_number
        sheet.cell(row=row, column=2).value = order.created_at.strftime("%d-%m-%Y")
        sheet.cell(row=row, column=3).value = order.status
        sheet.cell(row=row, column=4).value = str(order.required_date)
        sheet.cell(row=row, column=5).value = float(order.total_amount)

        total_purchase += float(order.total_amount)

        row += 1

    sheet.cell(row=row + 2, column=1).value = "Total Orders"
    sheet.cell(row=row + 2, column=2).value = orders.count()

    sheet.cell(row=row + 3, column=1).value = "Total Purchase"
    sheet.cell(row=row + 3, column=2).value = total_purchase

    return workbook